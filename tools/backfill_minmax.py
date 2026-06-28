#!/usr/bin/env python3
"""
Enkratni popravek: napolni prave dnevne MIN/MAX temperature v history.json.

Prvotni uvoz je v history.json prepisal samo dnevno povprečje ("Temperature")
v vsa tri polja (tempAvg = tempLow = tempHigh), zato so se v tabelah in grafih
prikazovale enake vrednosti za povprečje, minimum in maksimum.

Pravi vir je Ecowitt izvoz (all_Rečiškapstaja(...).xlsx), ki ima ločena
stolpca "Temperature Low(℃)" / "Temperature High(℃)" — za 2021-04 naprej.
Za zgodnje dneve (2019-11 … 2021-03), kjer postaja min/max ni beležila,
uporabimo Open-Meteo ERA5 reanalizo (enako kot rezervni vir v update_history.py).

    python3 tools/backfill_minmax.py            # zapiše spremembe
    python3 tools/backfill_minmax.py --dry-run  # samo poročilo

Posodobi se SAMO dneve, kjer je trenutno tempLow == tempHigh == tempAvg
(t.j. "sesedene" dneve). Prave meritve (src=station/era5) ostanejo nedotaknjene.
"""
import json, os, sys, glob, urllib.request, urllib.parse

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LAT, LON = 46.325779, 14.921137

DRY = "--dry-run" in sys.argv


def _num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def read_xlsx_minmax():
    """{date: {'avg':..,'low':..,'high':..}} iz Ecowitt izvoza (xlsx)."""
    import openpyxl
    matches = glob.glob(os.path.join(ROOT, "all_*.xlsx"))
    if not matches:
        sys.exit("Ne najdem Ecowitt izvoza (all_*.xlsx) v korenu projekta.")
    path = matches[0]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["result_list"] if "result_list" in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # skupinska glava
    next(rows, None)  # glava stolpcev
    # stolpci (0-osnovani): 0 Time, 1 Temperature, 2 Temperature Low, 3 Temperature High
    out = {}
    for r in rows:
        if not r or not r[0]:
            continue
        date = str(r[0])[:10]
        avg, low, high = _num(r[1]), _num(r[2]), _num(r[3])
        out[date] = {"avg": avg, "low": low, "high": high}
    return out


def fetch_era5(start, end):
    """{date: {'avg':..,'low':..,'high':..}} iz Open-Meteo ERA5 reanalize."""
    params = urllib.parse.urlencode({
        "latitude": LAT, "longitude": LON,
        "start_date": start, "end_date": end,
        "daily": "temperature_2m_max,temperature_2m_min,temperature_2m_mean",
        "timezone": "Europe/Berlin",
    })
    url = "https://archive-api.open-meteo.com/v1/archive?" + params
    with urllib.request.urlopen(url, timeout=90) as r:
        j = json.load(r)
    d = j.get("daily", {})
    out = {}
    for i, date in enumerate(d.get("time", [])):
        out[date] = {
            "avg":  _num(d.get("temperature_2m_mean", [None] * (i + 1))[i]),
            "low":  _num(d.get("temperature_2m_min",  [None] * (i + 1))[i]),
            "high": _num(d.get("temperature_2m_max",  [None] * (i + 1))[i]),
        }
    return out


def is_collapsed(v):
    lo, hi, av = v.get("tempLow"), v.get("tempHigh"), v.get("tempAvg")
    return lo is not None and lo == hi == av


def main():
    hp = os.path.join(ROOT, "history.json")
    hist = json.load(open(hp, encoding="utf-8"))

    collapsed = sorted(d for d, v in hist.items() if is_collapsed(v))
    print(f"Sesedenih dni (tempLow==tempHigh==tempAvg): {len(collapsed)}")

    xlsx = read_xlsx_minmax()

    n_station = 0          # popravljeno iz pravih meritev (xlsx low/high)
    need_era5 = []         # zgodnji dnevi brez min/max v izvozu
    for date in collapsed:
        x = xlsx.get(date)
        if x and x["low"] is not None and x["high"] is not None:
            v = hist[date]
            v["tempLow"]  = round(x["low"], 1)
            v["tempHigh"] = round(x["high"], 1)
            if x["avg"] is not None:
                v["tempAvg"] = round(x["avg"], 1)
            v.setdefault("src", "station")
            n_station += 1
        else:
            need_era5.append(date)

    n_era5 = 0
    if need_era5:
        lo, hi = min(need_era5), max(need_era5)
        print(f"Zgodnjih dni brez min/max v izvozu: {len(need_era5)} "
              f"({lo} … {hi}) → ERA5 reanaliza")
        era5 = fetch_era5(lo, hi)
        for date in need_era5:
            e = era5.get(date)
            if e and e["low"] is not None and e["high"] is not None:
                v = hist[date]
                v["tempLow"]  = round(e["low"], 1)
                v["tempHigh"] = round(e["high"], 1)
                if e["avg"] is not None:
                    v["tempAvg"] = round(e["avg"], 1)
                v["src"] = "era5"
                n_era5 += 1

    still = [d for d in collapsed if is_collapsed(hist[d])]
    print(f"Popravljeno iz postaje (xlsx low/high): {n_station}")
    print(f"Popravljeno iz ERA5:                    {n_era5}")
    print(f"Še vedno sesedeno po popravku:          {len(still)}")

    if DRY:
        print("\n--dry-run: history.json NI bil spremenjen.")
        return

    out = {k: hist[k] for k in sorted(hist)}
    json.dump(out, open(hp, "w", encoding="utf-8"),
              ensure_ascii=False, separators=(",", ":"))
    print(f"\n✓ history.json zapisan ({len(out)} dni).")


if __name__ == "__main__":
    main()
